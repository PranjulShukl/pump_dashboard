from django.shortcuts import render
from django.http import JsonResponse
from my_app.models import Inverter1Data
from .forms import ReportForm
from openpyxl.styles import Font, Alignment
from datetime import timedelta
from django.http import HttpResponse
from my_app.models import Inverter1Data
import pandas as pd


# Create your views here.
from django.shortcuts import render

# Home page view
def home(request):
    return render(request, 'home.html')

def about(request):
    return render(request, 'about.html')

def report(request):
    return render(request, 'report.html')

def evaluation(request):
    return render(request, 'evaluation.html')

def inverter1(request):
    return render(request, 'inverter1.html')

def weather(request):
    return render(request, 'weather.html')


def api_inverter1_data(request):
    inverter1_data = Inverter1Data.objects.last()
    data = {
        'holding_reg1': inverter1_data.holding_reg1 if inverter1_data else 0,
        'holding_reg2': inverter1_data.holding_reg2 if inverter1_data else 0,
        'holding_reg3': inverter1_data.holding_reg3 if inverter1_data else 0,
        'holding_reg4': inverter1_data.holding_reg4 if inverter1_data else 0,
        'holding_reg5': inverter1_data.holding_reg5 if inverter1_data else 0,
        'holding_reg6': inverter1_data.holding_reg6 if inverter1_data else 0,
        'holding_reg7': inverter1_data.holding_reg7 if inverter1_data else 0,
        'holding_reg8': inverter1_data.holding_reg8 if inverter1_data else 0,
        'holding_reg9': inverter1_data.holding_reg9 if inverter1_data else 0,
        'holding_reg10': inverter1_data.holding_reg10 if inverter1_data else 0,
        'holding_reg11': inverter1_data.holding_reg11 if inverter1_data else 0,
        'holding_reg12': inverter1_data.holding_reg12 if inverter1_data else 0,
        'last_updated': inverter1_data.last_updated if inverter1_data else 0,
    }
    return JsonResponse(data)


def report(request):
    sampled_data = []  # Initialize sampled_data to avoid errors if not POST

    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            # Get data from form
            start_time = form.cleaned_data['start_time']
            end_time = form.cleaned_data['end_time']
            interval_type = form.cleaned_data['interval_type']
            interval_duration = form.cleaned_data['interval_duration']

            # Map interval type to timedelta arguments
            interval_mapping = {
                'minute': timedelta(minutes=interval_duration),
                'hour': timedelta(hours=interval_duration),
                'day': timedelta(days=interval_duration),
                'month': timedelta(days=30 * interval_duration),  # Approximation for months
            }
            interval_timedelta = interval_mapping[interval_type]

            # Query SensorData model for the time range
            sensor_data = Inverter1Data.objects.filter(last_updated__range=(start_time, end_time))

            # Sample data at specific intervals
            current_time = start_time
            last_updated_time = None
            while current_time <= end_time:
                entry = sensor_data.filter(last_updated=current_time).first()
                if not entry:
                    entry = sensor_data.filter(last_updated__lte=current_time).order_by('-last_updated').first()

                if entry and ((last_updated_time is None) or (entry.last_updated != last_updated_time)):
                    last_updated_time = entry.last_updated  # Update last updated time
                    sampled_data.append(entry)  # Add to sampled data if exists

                current_time += interval_timedelta  # Move to next interval

            # Handle case where no data was sampled
            if not sampled_data:
                return HttpResponse("No data found for the given time range.", status=404)

            # If "Download Report" button was clicked
            if 'download' in request.POST:
                # Prepare the DataFrame
                data = {
                    'Last Updated': [
                        entry.last_updated.astimezone().replace(tzinfo=None) for entry in sampled_data
                    ],
                    'CH4': [entry.holding_reg1 for entry in sampled_data],
                    'CO2': [entry.holding_reg2 for entry in sampled_data],
                    'O2': [entry.holding_reg3 for entry in sampled_data],
                    'CV': [entry.holding_reg4 for entry in sampled_data],
                    'GCV': [entry.holding_reg5 for entry in sampled_data],
                    'NCV': [entry.holding_reg6 for entry in sampled_data],
                    'BALANCE': [entry.holding_reg7 for entry in sampled_data],
                    'FLOW VB': [entry.holding_reg8 for entry in sampled_data],
                    'FLOW VM': [entry.holding_reg9 for entry in sampled_data],
                    'PRESSURE': [entry.holding_reg10 for entry in sampled_data],
                    'TEMPERATURE': [entry.holding_reg11 for entry in sampled_data],
                    'cvg_volve_status': [entry.holding_reg12 for entry in sampled_data],
                }
                df = pd.DataFrame(data)

                # Create an Excel response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="sensor_data_report.xlsx"'

                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, startrow=4, sheet_name='Sensor Data')

                    # Access the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Sensor Data']

                    # Header styling
                    header_font = Font(bold=True, size=12)
                    header_alignment = Alignment(horizontal='center', vertical='center')

                    # Write and style the header information in the first rows
                    worksheet.merge_cells('A1:L1')
                    worksheet.cell(row=1, column=1, value=f"Device_name: Biospark1").font = header_font
                    worksheet.cell(row=1, column=1).alignment = header_alignment

                    worksheet.merge_cells('A2:L2')
                    worksheet.cell(row=2, column=1, value=f"Start Date: {start_time}").font = header_font
                    worksheet.cell(row=2, column=1).alignment = header_alignment

                    worksheet.merge_cells('A3:L3')
                    worksheet.cell(row=3, column=1, value=f"End Date: {end_time}").font = header_font
                    worksheet.cell(row=3, column=1).alignment = header_alignment

                return response

    else:
        form = ReportForm()

    return render(request, 'report.html', {'form': form, 'sampled_data': sampled_data})
